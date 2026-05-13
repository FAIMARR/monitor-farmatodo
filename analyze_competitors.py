import re
import openpyxl
from pathlib import Path
from datetime import datetime

def extract_presentation(name):
    """
    Extrae la presentación (volumen, peso, unidades) de un nombre de producto.
    Ejemplo: 'Crema Facial Nivea 50ml' -> '50ml'
    """
    if not isinstance(name, str):
        return "N/A"
    
    # Patrones comunes: 50ml, 100 g, 30 caps, 2 unid, etc.
    pattern = r'(\d+(?:[\.,]\d+)?\s*(?:ml|g|und|caps|tab|mg|oz|l|cc|gr|unid|tabletas|cápsulas|ml\.|gr\.|uds))'
    match = re.search(pattern, name, re.IGNORECASE)
    if match:
        return match.group(1).lower().replace(" ", "")
    return "N/A"

def clean_product_type(name, brand, subcategory):
    """
    Intenta extraer el 'tipo' de producto eliminando la marca y la presentación.
    """
    if not isinstance(name, str):
        return "Desconocido"
    
    name_clean = name.lower()
    if isinstance(brand, str):
        name_clean = name_clean.replace(brand.lower(), "").strip()
    
    presentation = extract_presentation(name)
    if presentation != "N/A":
        name_clean = name_clean.replace(presentation, "").strip()
    
    name_clean = re.sub(r'[\-\s,\.]+$', '', name_clean)
    name_clean = re.sub(r'^[\-\s,\.]+', '', name_clean)
    
    if len(name_clean) < 3:
        return subcategory if isinstance(subcategory, str) else "General"
    
    return name_clean.title()

def analyze_excel(file_path):
    print(f"Analizando: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Mapeo de cabeceras
    headers = [cell.value for cell in ws[1]]
    try:
        idx_prod = headers.index("Producto")
        idx_brand = headers.index("Marca")
        idx_subcat = headers.index("Subcategoria")
    except ValueError:
        print("Error: No se encontraron las columnas necesarias (Producto, Marca, Subcategoria)")
        return

    # Añadir nuevas cabeceras
    new_headers = ["Presentación", "Tipo Base", "Grupo Competencia"]
    for i, h in enumerate(new_headers, 1):
        ws.cell(row=1, column=len(headers) + i, value=h)

    print("Procesando filas...")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for r_idx, row in enumerate(rows, 2):
        name = row[idx_prod]
        brand = row[idx_brand]
        subcat = row[idx_subcat]
        
        pres = extract_presentation(name)
        base_type = clean_product_type(name, brand, subcat)
        group_id = f"{subcat} | {pres}"
        
        ws.cell(row=r_idx, column=len(headers) + 1, value=pres)
        ws.cell(row=r_idx, column=len(headers) + 2, value=base_type)
        ws.cell(row=r_idx, column=len(headers) + 3, value=group_id)

    output_path = file_path.parent / f"COMPETIDORES_{file_path.name}"
    wb.save(output_path)
    print(f"¡Listo! Archivo de competidores guardado: {output_path}")

if __name__ == "__main__":
    export_dir = Path("exports")
    if not export_dir.exists():
        print("La carpeta 'exports' no existe.")
    else:
        files = list(export_dir.glob("Farmatodo_*.xlsx"))
        if not files:
            print("No se encontraron archivos .xlsx en 'exports'.")
        else:
            latest_file = max(files, key=lambda x: x.stat().st_mtime)
            analyze_excel(latest_file)
