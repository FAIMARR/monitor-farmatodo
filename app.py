# -*- coding: utf-8 -*-
"""
Farmatodo Venezuela - Web App de Búsqueda de Productos
Backend: Flask + Playwright + openpyxl
"""
import sys, io, json, asyncio, threading, time, re, os
from datetime import datetime
from pathlib import Path
from queue import Queue, Empty

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from flask import Flask, Response, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from playwright.async_api import async_playwright, TimeoutError as PWT

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

@app.route("/")
def serve_index():
    return send_from_directory('.', 'index.html')

@app.route("/style.css")
def serve_css():
    return send_from_directory('.', 'style.css', mimetype='text/css')

@app.route("/app.js")
def serve_js():
    return send_from_directory('.', 'app.js', mimetype='application/javascript')

BASE_URL = "https://www.farmatodo.com.ve"
EXPORT_DIR = Path(__file__).parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

CATEGORIES = {
    "Salud y Medicamentos": [
        f"{BASE_URL}/categorias/salud-y-medicamentos/medicamentos",
        f"{BASE_URL}/categorias/salud-y-medicamentos/vitaminas-y-productos-naturales",
        f"{BASE_URL}/categorias/salud-y-medicamentos/dolor-general",
        f"{BASE_URL}/categorias/salud-y-medicamentos/salud-digestiva",
        f"{BASE_URL}/categorias/salud-y-medicamentos/salud-respiratoria-y-gripe",
        f"{BASE_URL}/categorias/salud-y-medicamentos/nutricion-y-vida-saludable",
        f"{BASE_URL}/categorias/salud-y-medicamentos/botiquin-y-primeros-auxilios",
        f"{BASE_URL}/categorias/salud-y-medicamentos/cuidado-de-la-vista",
        f"{BASE_URL}/categorias/salud-y-medicamentos/dermatologicos",
    ],
    "Cuidado Personal": [
        f"{BASE_URL}/categorias/cuidado-personal/cuidado-del-cabello",
        f"{BASE_URL}/categorias/cuidado-personal/cuidado-de-la-piel",
        f"{BASE_URL}/categorias/cuidado-personal/higiene-personal",
        f"{BASE_URL}/categorias/cuidado-personal/cuidado-bucal",
        f"{BASE_URL}/categorias/cuidado-personal/afeitado",
        f"{BASE_URL}/categorias/cuidado-personal/proteccion-femenina",
    ],
    "Belleza": [
        f"{BASE_URL}/categorias/belleza/cosmeticos",
        f"{BASE_URL}/categorias/belleza/fragancias",
        f"{BASE_URL}/categorias/belleza/artefactos-de-belleza",
        f"{BASE_URL}/categorias/belleza/mercancia-general",
    ],
    "Bebé": [
        f"{BASE_URL}/categorias/bebe/higiene-del-bebe",
        f"{BASE_URL}/categorias/bebe/bebe-alimentos",
    ],
    "Alimentos y Bebidas": [
        f"{BASE_URL}/categorias/alimentos-y-bebidas/alimentos",
        f"{BASE_URL}/categorias/alimentos-y-bebidas/bebidas",
        f"{BASE_URL}/categorias/alimentos-y-bebidas/dulces-y-snacks",
    ],
    "Hogar y Mascotas": [
        f"{BASE_URL}/categorias/hogar-mascota-y-otros/higiene-del-hogar",
        f"{BASE_URL}/categorias/hogar-mascota-y-otros/decoracion-util-hogar",
        f"{BASE_URL}/categorias/hogar-mascota-y-otros/tecnologia",
    ],
}

# JavaScript de extracción (selectores reales de Farmatodo)
EXTRACT_JS = """
() => {
    const results = [];
    const seen = new Set();

    // Metodo 1: selector nativo confirmado a.product-card__info-link
    const infoLinks = document.querySelectorAll('a.product-card__info-link');
    infoLinks.forEach(link => {
        const href = link.href || '';
        if (seen.has(href)) return;
        seen.add(href);
        const paras = link.querySelectorAll('p');
        const spans = link.querySelectorAll('span');
        const card  = link.closest('ftd-card-product, [class*="card-product"], [class*="product-card"]');
        const discEl = card ? card.querySelector('[class*="discount"],[class*="badge"],[class*="-off"],[class*="porcent"]') : null;
        const imgEl  = card ? card.querySelector('img[src]') : null;
        const brand  = paras[0] ? paras[0].innerText.trim() : '';
        const name   = paras[1] ? paras[1].innerText.trim() : (paras[0] ? paras[0].innerText.trim() : '');
        const price  = spans[0] ? spans[0].innerText.trim() : '';
        const oldPr  = spans[1] ? spans[1].innerText.trim() : '';
        const disc   = discEl  ? discEl.innerText.trim()    : '';
        const img    = imgEl   ? imgEl.src : '';
        if (name || price) results.push({name, brand, price, oldPrice: oldPr, discount: disc, link: href, image: img});
    });

    // Metodo 2: fallback cards genericas
    if (results.length === 0) {
        const sels = ['ftd-card-product','app-card-product','[class*="product-card"]','[class*="card-product"]'];
        let cards = [];
        for (const s of sels) { const f = [...document.querySelectorAll(s)]; if (f.length) { cards = f; break; } }
        cards.forEach(card => {
            const a = card.querySelector('a[href]');
            const href = a ? a.href : '';
            if (seen.has(href)) return;
            seen.add(href);
            const ps = [...card.querySelectorAll('p,h3,h4')];
            const brand = ps[0] ? ps[0].innerText.trim() : '';
            const name  = ps[1] ? ps[1].innerText.trim() : brand;
            const sps   = [...card.querySelectorAll('span')].filter(s => /[0-9]/.test(s.innerText));
            const price   = sps[0] ? sps[0].innerText.trim() : '';
            const oldPrice = sps[1] ? sps[1].innerText.trim() : '';
            const img = (card.querySelector('img') || {src:''}).src;
            if (name || price) results.push({name, brand, price, oldPrice, discount:'', link: href, image: img});
        });
    }
    return results;
}
"""

# ── Estado global de sesiones de scraping ──────────────────────────
_sessions: dict[str, dict] = {}


def _new_event_loop():
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    return loop


async def _dismiss(page):
    for sel in ['button:has-text("Entendido")','button:has-text("Aceptar")',
                'button:has-text("Continuar")','button:has-text("Cerrar")',
                '[aria-label="Close"]','[aria-label="close"]']:
        try:
            b = page.locator(sel).first
            if await b.is_visible(timeout=1200): await b.click(); await asyncio.sleep(0.6)
        except Exception: pass


def _parse_product(raw: dict, subcat: str, search_query: str = "") -> dict | None:
    """
    Normaliza un producto crudo desde la API de Farmatodo o Algolia
    a nuestro formato estándar. Retorna None si no pasa el filtro de búsqueda.
    """
    # ── Formato Algolia (hits[]) ──────────────────────────────────
    name  = (raw.get("name") or raw.get("productName") or
             raw.get("nombre") or raw.get("displayName") or "")
    brand = (raw.get("brand") or raw.get("brandName") or
             raw.get("marca") or raw.get("laboratory") or "")
    link  = raw.get("url") or raw.get("productUrl") or raw.get("link") or ""
    image = (raw.get("image") or raw.get("imageUrl") or
             raw.get("thumbnail") or raw.get("img") or "")

    # Precio — puede venir en distintos formatos
    price_raw = (raw.get("price") or raw.get("precio") or
                 raw.get("salePrice") or raw.get("currentPrice") or 0)
    if isinstance(price_raw, dict):
        price_val = price_raw.get("value") or price_raw.get("amount") or price_raw.get("sale") or 0
    else:
        price_val = price_raw

    old_raw = (raw.get("originalPrice") or raw.get("listPrice") or
               raw.get("normalPrice") or raw.get("compareAtPrice") or 0)
    if isinstance(old_raw, dict):
        old_val = old_raw.get("value") or old_raw.get("amount") or 0
    else:
        old_val = old_raw

    disc = raw.get("discount") or raw.get("discountPercentage") or ""
    if isinstance(disc, (int, float)) and disc:
        disc = f"-{int(disc)}%"

    # Formatear precio en Bolívares
    def fmt_price(v):
        try:
            f = float(v)
            return f"Bs.{f:,.2f}" if f > 0 else ""
        except Exception:
            return str(v) if v else ""

    # Construir URL completa
    if link and not link.startswith("http"):
        link = BASE_URL + ("/" if not link.startswith("/") else "") + link
    if not link and raw.get("objectID"):
        link = f"{BASE_URL}/producto/{raw['objectID']}"

    # Filtro de búsqueda (si aplica)
    if search_query:
        q = search_query.lower()
        if q not in name.lower() and q not in brand.lower():
            return None

    if not name and not price_val:
        return None

    # Extracción y limpieza de SKU
    sku = str(raw.get("sku") or raw.get("item") or raw.get("itemId") or
              raw.get("productId") or raw.get("id") or raw.get("objectID") or "")
    
    if not sku and link:
        m = re.search(r'(?:producto|product|p)/(\d+)', link) or re.search(r'(\d{6,14})', link)
        if m: sku = m.group(1)
    
    if not sku and image:
        m = re.search(r'(\d{6,14})', image)
        if m: sku = m.group(1)

    # Eliminar cualquier comilla prefijada
    sku = sku.replace("'", "").replace('"', "").strip()

    return {
        "name":       name,
        "brand":      brand,
        "price":      fmt_price(price_val),
        "oldPrice":   fmt_price(old_val),
        "discount":   str(disc) if disc else "",
        "link":       link,
        "image":      image,
        "subcategory": subcat,
        "sku":        sku,
    }

# ── Funciones de Análisis de Competencia ──────────────────────

def _extract_presentation(name):
    """Extrae volumen/peso del nombre del producto."""
    if not name: return "N/A"
    pattern = r'(\d+(?:[\.,]\d+)?\s*(?:ml|g|und|caps|tab|mg|oz|l|cc|gr|unid|tabletas|cápsulas|ml\.|gr\.|uds))'
    match = re.search(pattern, name, re.IGNORECASE)
    return match.group(1).lower().replace(" ", "") if match else "N/A"

def _clean_product_type(name, brand, subcat):
    """Limpia el nombre para obtener el tipo base del producto."""
    if not name: return "General"
    n = name.lower()
    if brand: n = n.replace(brand.lower(), "").strip()
    pres = _extract_presentation(name)
    if pres != "N/A": n = n.replace(pres, "").strip()
    n = re.sub(r'[\-\s,\.]+$', '', n)
    n = re.sub(r'^[\-\s,\.]+', '', n)
    return n.title() if len(n) > 2 else (subcat or "General")


# ── Notificaciones de Correo (Gratis vía SMTP) ──────────────────────────────

EMAIL_CONFIG_FILE = Path(__file__).parent / "email_config.json"

def _load_email_config():
    return _load_json(EMAIL_CONFIG_FILE, {
        "smtp_server": "smtp.gmail.com",
        "smtp_port": 587,
        "sender_email": "",
        "sender_password": "", # Usar 'App Password' en Gmail
        "receiver_email": "",
        "enabled": False
    })

def _send_email_notification(subject, html_content):
    cfg = _load_email_config()
    if not cfg.get("enabled"): return False
    
    # Procesar múltiples destinatarios (separados por comas)
    receivers_raw = cfg.get("receiver_email", "")
    receivers = [r.strip() for r in receivers_raw.split(",") if "@" in r]
    
    if not receivers: 
        print("[Email] No hay destinatarios válidos.")
        return False

    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        msg = MIMEMultipart()
        msg['From'] = cfg.get("sender_email")
        msg['To'] = ", ".join(receivers)
        msg['Subject'] = subject

        msg.attach(MIMEText(html_content, 'html'))

        server = smtplib.SMTP(cfg['smtp_server'], cfg['smtp_port'])
        server.starttls()
        server.login(cfg['sender_email'], cfg['sender_password'])
        server.send_message(msg)
        server.quit()
        print(f"[Email] Alerta enviada a {cfg['receiver_email']}")
    except Exception as e:
        print(f"[Email] Error enviando correo: {e}")


# ── Persistencia en la Nube (vía Google Sheets) ──────────────────────────

def _sync_state_to_sheets():
    """Guarda la Watchlist y Config en una pestaña oculta de Google Sheets."""
    try:
        cfg_gs = _load_json(GSHEETS_CONFIG_FILE, {})
        if not cfg_gs.get("spreadsheet_id"): return
        
        gc = _get_gsheets_client()
        sh = gc.open_by_key(cfg_gs["spreadsheet_id"])
        
        # Intentar abrir o crear pestaña de respaldo
        try:
            ws = sh.worksheet("SISTEMA_BACKUP")
        except:
            ws = sh.add_worksheet(title="SISTEMA_BACKUP", rows="100", cols="2")
        
        data = {
            "watchlist": _load_json(WATCHLIST_FILE, []),
            "email_config": _load_json(EMAIL_CONFIG_FILE, {})
        }
        ws.update("A1", [["DATA_JSON"]])
        ws.update("A2", [[json.dumps(data)]])
        print("[Cloud] Respaldo guardado en Google Sheets")
    except Exception as e:
        print(f"[Cloud] Error en respaldo: {e}")

def _recover_state_from_sheets():
    """Recupera el estado al arrancar si los archivos locales están vacíos."""
    try:
        cfg_gs = _load_json(GSHEETS_CONFIG_FILE, {})
        if not cfg_gs.get("spreadsheet_id"): return
        
        gc = _get_gsheets_client()
        sh = gc.open_by_key(cfg_gs["spreadsheet_id"])
        ws = sh.worksheet("SISTEMA_BACKUP")
        val = ws.acell("A2").value
        if val:
            data = json.loads(val)
            _save_json(WATCHLIST_FILE, data.get("watchlist", []))
            _save_json(EMAIL_CONFIG_FILE, data.get("email_config", {}))
            print("[Cloud] Estado recuperado desde Google Sheets")
    except:
        print("[Cloud] No se encontró respaldo para recuperar")

# ── Sistema de Vigilancia de Precios ───────────────────────────────────────

PRICE_DB_FILE = Path(__file__).parent / "price_history.json"
ALERTS_FILE   = Path(__file__).parent / "price_alerts.json"
WATCHLIST_FILE = Path(__file__).parent / "watchlist_backend.json"

def _load_json(path, default):
    if Path(path).exists():
        try: return json.loads(Path(path).read_text(encoding="utf-8"))
        except: return default
    return default

def _save_json(path, data):
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

# Estado del monitor
_monitor_active = False

def _parse_price_float(price_str):
    if not price_str: return 0.0
    try:
        clean = re.sub(r'[^\d,.]', '', price_str)
        if ',' in clean and '.' in clean:
            clean = clean.replace('.', '').replace(',', '.')
        elif ',' in clean:
            clean = clean.replace(',', '.')
        return float(clean)
    except:
        return 0.0

def _monitor_loop_sync():
    """Bucle de monitoreo compatible con Windows y Playwright."""
    global _monitor_active
    print("\n" + "!"*50)
    print("  [SISTEMA] MONITOR DE PRECIOS INICIADO")
    print("!"*50 + "\n")
    
    # En Windows, Playwright requiere ProactorEventLoop
    if os.name == 'nt':
        loop = asyncio.ProactorEventLoop()
    else:
        loop = asyncio.new_event_loop()
        
    asyncio.set_event_loop(loop)

    while _monitor_active:
        try:
            watchlist = _load_json(WATCHLIST_FILE, [])
            if not watchlist:
                time.sleep(30)
                continue

            history = _load_json(PRICE_DB_FILE, {})
            alerts  = _load_json(ALERTS_FILE, [])
            
            for prod in watchlist:
                if not _monitor_active: break
                url = prod.get("link")
                if not url: continue

                # Ejecutamos el scrape unitario de forma síncrona dentro del loop
                async def check_single(u):
                    async with async_playwright() as p:
                        browser = await p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-setuid-sandbox"])
                        context = await browser.new_context(
                            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
                        )
                        page = await context.new_page()
                        try:
                            await page.goto(u, timeout=45000, wait_until="networkidle")
                            await asyncio.sleep(2)
                            data = await page.evaluate(EXTRACT_JS)
                            return data[0] if data else None
                        except: return None
                        finally: await browser.close()

                current = loop.run_until_complete(check_single(url))
                if current:
                    new_price_str = current.get("price", "")
                    new_price_val = _parse_price_float(new_price_str)
                    sku = current.get("sku") or url
                    old_record = history.get(sku)

                    if old_record:
                        old_val = old_record.get("price_val", 0.0)
                        if new_price_val != old_val and new_price_val > 0:
                            diff = new_price_val - old_val
                            alert = {
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "name": current.get("name"),
                                "old_price": old_record.get("price_str"),
                                "new_price": new_price_str,
                                "diff": round(diff, 2),
                                "type": "UP" if diff > 0 else "DOWN",
                                "link": url,
                                "image": current.get("image")
                            }
                            alerts.insert(0, alert)
                            _save_json(ALERTS_FILE, alerts[:100])
                            
                            # ENVÍO DE EMAIL (Gratis)
                            emoji = "📈" if diff > 0 else "📉"
                            subject = f"{emoji} Alerta de Precio: {current.get('name')}"
                            
                            html = f"""
                            <div style="font-family: sans-serif; max-width: 600px; border: 1px solid #eee; padding: 20px; border-radius: 12px;">
                                <h2 style="color: {'#ff3d71' if diff > 0 else '#00c853'};">Cambio de Precio Detectado</h2>
                                <p>Hola, el producto <strong>{current.get('name')}</strong> ha cambiado de precio:</p>
                                <div style="background: #f9f9f9; padding: 15px; border-radius: 8px; margin: 20px 0;">
                                    <p>💰 Precio Anterior: <span style="text-decoration: line-through; color: #888;">{old_record.get('price_str')}</span></p>
                                    <p>💵 <strong>Precio Nuevo: {new_price_str}</strong></p>
                                    <p>📊 Variación: <strong>{round(diff, 2)}</strong></p>
                                </div>
                                <a href="{url}" style="display: inline-block; background: #2e6db4; color: white; padding: 12px 25px; text-decoration: none; border-radius: 6px; font-weight: bold;">Ver en Farmatodo</a>
                            </div>
                            """
                            _send_email_notification(subject, html)
                    
                    history[sku] = {
                        "name": current.get("name"), "price_str": new_price_str,
                        "price_val": new_price_val, "last_seen": datetime.now().strftime("%H:%M:%S")
                    }
                    _save_json(PRICE_DB_FILE, history)

                time.sleep(10) # Pausa entre productos

            # Esperar 30 min
            for _ in range(180):
                if not _monitor_active: break
                time.sleep(10)

        except Exception as e:
            print(f"[Monitor] Error general: {e}")
            time.sleep(60)

def _start_monitor_thread():
    global _monitor_active
    _monitor_active = True
    t = threading.Thread(target=_monitor_loop_sync, daemon=True)
    t.start()


# ── Motor de IA para análisis de competencia ───────────────────────────────

AI_CONFIG_FILE = Path(__file__).parent / "ai_config.json"

def _load_ai_config():
    if AI_CONFIG_FILE.exists():
        return json.loads(AI_CONFIG_FILE.read_text(encoding="utf-8"))
    return {"gemini_api_key": ""}

def _save_ai_config(cfg: dict):
    AI_CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def _ai_classify_batch(products_batch: list[dict], api_key: str) -> list[dict]:
    """
    Envía un lote de productos a Gemini para clasificación semántica de competidores.
    Retorna lista de dicts con: idx, product_type, use, target, competitor_group
    """
    lines = []
    for i, p in enumerate(products_batch, 1):
        brand = p.get("brand", "") or ""
        name  = p.get("name",  "") or ""
        subcat = p.get("subcategory", "") or ""
        lines.append(f"{i}. [Marca: {brand}] {name} [Categoría: {subcat}]")

    products_text = "\n".join(lines)

    prompt = f"""Eres un experto en análisis competitivo de mercado farmacéutico, cosmético y cuidado personal en Venezuela.

Tu tarea es clasificar cada producto con MÁXIMA PRECISIÓN para identificar competidores directos.

⚠️ REGLA FUNDAMENTAL: Dos productos son competidores DIRECTOS SOLO SI un consumidor podría comprar uno EN LUGAR del otro para el MISMO uso específico. Productos de la misma categoría general NO son necesariamente competidores.

Para cada producto determina:

1. **product_type**: Nombre ESPECÍFICO del tipo de producto. Debe ser tan preciso como sea posible.
   ✅ CORRECTO: "Lápiz Delineador de Ojos", "Lápiz/Lápiz Definidor de Cejas", "Brillo Labial", "Prebase/Primer de Maquillaje", "Máscara de Pestañas"
   ❌ INCORRECTO: "Cosmético", "Maquillaje", "Producto de belleza"

2. **use**: Función específica del producto en máximo 6 palabras.
   ✅ CORRECTO: "Delinear y definir contorno de ojos", "Rellenar y definir cejas", "Hidratar y dar brillo a labios"
   ❌ INCORRECTO: "Uso cosmético", "Para maquillaje"

3. **target**: A quién va dirigido (ej: "Mujeres adultas", "Piel grasa", "Bebés 0-3 años", "Adultos mayores", "Hombres").

4. **competitor_group**: Clave ULTRA-ESPECÍFICA que agrupa SOLO productos que el consumidor intercambiaría entre sí.
   
   TAXONOMÍA OBLIGATORIA — usa el formato: "Función Exacta | Formato/Presentación"
   
   EJEMPLOS CORRECTOS de productos cosméticos:
   - Delineador de ojos líquido negro → "Delineador Ojos Líquido Negro"
   - Lápiz de cejas → "Definidor/Lápiz de Cejas | 1g"  
   - Brillo labial → "Brillo Labial | 5ml"
   - Prebase de maquillaje → "Primer/Prebase Maquillaje | 30-40ml"
   - Máscara de pestañas negra → "Máscara Pestañas Negra"
   - Rubor en polvo → "Rubor/Colorete Polvo"
   - Sombra de ojos → "Sombra de Ojos Paleta"
   - Base líquida de maquillaje → "Base Maquillaje Líquida | 30ml"
   - Corrector de ojeras → "Corrector/Concealer Facial"
   
   EJEMPLOS CORRECTOS de productos farmacéuticos/cuidado personal:
   - Paracetamol 500mg tabletas → "Paracetamol Oral | 500mg Tab"
   - Ibuprofeno 400mg → "Ibuprofeno Oral | 400mg"
   - Crema hidratante facial → "Crema Hidratante Facial | 50ml"
   - Champú anticaspa → "Champú Anticaspa Control Grasa"
   - Bloqueador solar SPF 50 facial → "Protector Solar Facial SPF50"
   - Agua micelar desmaquillante → "Agua Micelar Desmaquillante | 400ml"
   - Gel de ducha corporal → "Gel/Jabón Ducha Corporal | 250ml"
   
   REGLAS ESTRICTAS:
   ✅ Mismo tipo EXACTO de producto (lápiz de cejas ≠ delineador de ojos ≠ brillo labial)
   ✅ Misma función principal (hidratación ≠ nutrición ≠ control grasa)  
   ✅ Misma zona de aplicación (labios ≠ ojos ≠ cejas ≠ rostro ≠ cuerpo)
   ✅ Presentación similar (variación ±30%% aceptable: 48ml y 50ml son iguales, 50ml y 200ml son distintos)
   ❌ NUNCA agrupar solo por marca, categoría general o subcategoría
   ❌ NUNCA agrupar "Delineador de ojos" con "Lápiz de cejas" aunque ambos sean "productos para ojos"
   ❌ NUNCA agrupar "Brillo labial" con "Labial mate" aunque ambos sean "productos para labios"

Responde ÚNICAMENTE con un JSON array válido, sin texto adicional, sin markdown:
[{{"idx": 1, "product_type": "...", "use": "...", "target": "...", "competitor_group": "..."}}, ...]

Productos a analizar:
{products_text}"""

    try:
        from google import genai
        from google.genai import types

        client = genai.Client(api_key=api_key)

        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0.1,
                max_output_tokens=8192,
            ),
        )
        raw = response.text.strip()
        # Limpiar posible markdown
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)
        result = json.loads(raw)
        return result if isinstance(result, list) else []
    except Exception as e:
        print(f"[AI] Error en lote: {e}")
        return []


# Estado global de análisis IA
_ai_sessions: dict[str, dict] = {}

def _run_ai_analysis(ai_session_id: str, products: list[dict], api_key: str):
    """
    Corre el análisis IA en un hilo separado.
    Emite progreso por una Queue y guarda resultados en _ai_sessions.
    """
    sess = _ai_sessions[ai_session_id]
    q: Queue = sess["queue"]
    BATCH = 25  # productos por llamada a Gemini

    total = len(products)
    enriched = [p.copy() for p in products]  # copia para no mutar originales

    try:
        for batch_start in range(0, total, BATCH):
            if sess.get("cancelled"):
                break
            batch      = products[batch_start:batch_start + BATCH]
            batch_end  = min(batch_start + BATCH, total)
            pct        = int((batch_start / total) * 90)

            q.put({"event": "ai_status", "data": {
                "msg": f"Analizando productos {batch_start+1}–{batch_end} de {total}...",
                "pct": pct
            }})

            classifications = _ai_classify_batch(batch, api_key)

            # Mapear resultados al array enriched
            for c in classifications:
                idx = c.get("idx", 0) - 1  # 1-indexed → 0-indexed
                abs_idx = batch_start + idx
                if 0 <= abs_idx < len(enriched):
                    enriched[abs_idx]["ai_product_type"]    = c.get("product_type", "")
                    enriched[abs_idx]["ai_use"]             = c.get("use", "")
                    enriched[abs_idx]["ai_target"]          = c.get("target", "")
                    enriched[abs_idx]["ai_competitor_group"]= c.get("competitor_group", "")
                    q.put({"event": "ai_product", "data": {
                        "abs_idx": abs_idx,
                        "ai_product_type":     c.get("product_type", ""),
                        "ai_use":              c.get("use", ""),
                        "ai_target":           c.get("target", ""),
                        "ai_competitor_group": c.get("competitor_group", ""),
                    }})

        sess["enriched_products"] = enriched
        q.put({"event": "ai_done", "data": {"total": total}})

    except Exception as e:
        q.put({"event": "ai_error", "data": {"msg": str(e)}})


async def _scrape(session_id: str, urls: list[str], search_query: str = ""):
    """
    Versión ULTRA-LIGERA para Render Free.
    Sin interceptación de red pesada, solo carga y extracción directa.
    """
    from playwright.async_api import async_playwright, TimeoutError as PWT
    sess = _sessions[session_id]
    q: Queue = sess["queue"]
    
    def emit(event: str, data: dict):
        q.put({"event": event, "data": data})

    all_products = []
    seen_keys = set()
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage", "--single-process"])
        context = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
        
        for url_idx, url in enumerate(urls):
            if sess.get("cancelled"): break
            page = await context.new_page()
            # Bloquear todo lo innecesario
            await page.route("**/*.{png,jpg,jpeg,gif,webp,svg,css,woff,woff2,map}", lambda r: r.abort())
            
            subcat = url.split("/")[-1].replace("-", " ").title()
            pct_base = int(5 + (url_idx / len(urls)) * 85)
            emit("status", {"msg": f"Buscando en {subcat}...", "pct": pct_base})
            
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                await asyncio.sleep(2)
                
                # Extraer datos usando el JS optimizado
                products = await page.evaluate(EXTRACT_JS)
                for p in products:
                    key = p.get("link") or p.get("name")
                    if key and key not in seen_keys:
                        p["timestamp"] = now_str
                        p["subcategory"] = subcat
                        seen_keys.add(key)
                        all_products.append(p)
                        emit("product", p)
                
            except Exception: pass
            finally: await page.close()

    sess["products"] = all_products
    emit("done", {"total": len(all_products)})


def _run_scrape(session_id, urls, search_query):
    loop = _new_event_loop()
    loop.run_until_complete(_scrape(session_id, urls, search_query))
    loop.close()


# ── Rutas ──────────────────────────────────────────────────────────

@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/api/categories")
def get_categories():
    return jsonify({k: len(v) for k, v in CATEGORIES.items()})


@app.route("/api/scrape", methods=["POST"])
def start_scrape():
    body      = request.get_json(force=True)
    category  = body.get("category", "")
    terms     = body.get("terms", [])       # lista de términos individuales
    session_id = f"sess_{int(time.time()*1000)}"

    # Normalizar: quitar vacíos
    terms = [t.strip() for t in terms if str(t).strip()]

    # Construir URLs:
    # - Si hay términos → una URL /buscar?q= por cada término (cubrimos todos)
    # - Si hay categoría → sumar sus subcategorías
    # - Si nada → todas las categorías
    urls = []
    if terms:
        urls = [f"{BASE_URL}/buscar?product={t}&departamento=Todos&filtros=" for t in terms]
    if category and category in CATEGORIES:
        urls += CATEGORIES[category]
    if not urls:
        urls = [u for sub in CATEGORIES.values() for u in sub]

    _sessions[session_id] = {
        "queue":    Queue(),
        "products": [],
        "cancelled": False,
        "terms":    terms,
    }

    # search_q se pasa vacío al scraper → NO filtra server-side, lo hace el frontend
    t = threading.Thread(target=_run_scrape, args=(session_id, urls, ""), daemon=True)
    t.start()
    return jsonify({"session_id": session_id})


@app.route("/api/events/<session_id>")
def sse_events(session_id):
    """Server-Sent Events — stream de progreso y productos en tiempo real."""
    def generate():
        if session_id not in _sessions:
            yield "data: {\"event\":\"error\",\"data\":{\"msg\":\"Sesion no encontrada\"}}\n\n"
            return
        q: Queue = _sessions[session_id]["queue"]
        while True:
            try:
                msg = q.get(timeout=30)
                yield f"data: {json.dumps(msg)}\n\n"
                if msg["event"] == "done":
                    break
            except Empty:
                yield "data: {\"event\":\"ping\"}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/cancel/<session_id>", methods=["POST"])
def cancel_scrape(session_id):
    if session_id in _sessions:
        _sessions[session_id]["cancelled"] = True
    return jsonify({"ok": True})


@app.route("/api/export/<session_id>")
def export_excel(session_id):
    if session_id not in _sessions:
        return jsonify({"error": "Sesion no encontrada"}), 404

    products = _sessions[session_id].get("products", [])
    if not products:
        return jsonify({"error": "Sin productos para exportar"}), 400

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

        HEADERS = ["N", "Fecha/Hora", "Subcategoria", "Marca", "Producto", "SKU", "Precio", "Precio Anterior", "Descuento", "URL", "Presentación", "Tipo Base", "Grupo Competencia"]
        WIDTHS  = [5,    20,           22,             20,      52,         15,    14,       16,                12,          55,    15,             25,          30]
        BLUE = "2E6DB4"
        ALT  = "EBF4FF"

        s = Side(style="thin", color="FFFFFF")
        for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
            c = ws.cell(1, ci, h)
            c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = Border(left=s, right=s, top=s, bottom=s)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 26

        ds = Side(style="thin", color="D0D0D0")
        for ri, p in enumerate(products, 2):
            name   = p.get("name", "")
            brand  = p.get("brand", "")
            subcat = p.get("subcategory", "")
            pres   = _extract_presentation(name)
            base   = _clean_product_type(name, brand, subcat)
            group  = f"{subcat} | {pres}"

            row = [ri-1, p.get("timestamp",""), subcat, brand,
                   name, p.get("sku",""), p.get("price",""), p.get("oldPrice",""),
                   p.get("discount",""), p.get("link",""), pres, base, group]
            bg = ALT if ri % 2 == 0 else "FFFFFF"
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.font      = Font(size=10, name="Calibri")
                c.alignment = Alignment(vertical="center")
                c.border    = Border(left=ds, right=ds, top=ds, bottom=ds)
            ws.row_dimensions[ri].height = 16

        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

        fname = f"Farmatodo_{session_id[-6:]}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        fpath = EXPORT_DIR / fname
        wb.save(fpath)
        return send_file(str(fpath), as_attachment=True, download_name=fname)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Google Sheets ──────────────────────────────────────────────
GSHEETS_CONFIG_FILE = Path(__file__).parent / "gsheets_config.json"

def _load_gsheets_config():
    if GSHEETS_CONFIG_FILE.exists():
        return json.loads(GSHEETS_CONFIG_FILE.read_text(encoding="utf-8"))
    return {"spreadsheet_url": "", "credentials_path": ""}

@app.route("/api/gsheets/config", methods=["GET"])
def get_gsheets_config():
    return jsonify(_load_gsheets_config())

@app.route("/api/gsheets/config", methods=["POST"])
def save_gsheets_config():
    cfg = request.get_json(force=True)
    GSHEETS_CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    return jsonify({"ok": True})

@app.route("/api/gsheets/export/<session_id>", methods=["POST"])
def export_to_gsheets(session_id):
    if session_id not in _sessions:
        return jsonify({"error": "Sesión no encontrada"}), 404

    products = _sessions[session_id].get("products", [])
    if not products:
        return jsonify({"error": "Sin productos para exportar"}), 400

    cfg = _load_gsheets_config()
    creds_path = cfg.get("credentials_path", "").strip().strip('"').strip("'")
    sheet_url  = cfg.get("spreadsheet_url", "").strip()

    if not creds_path:
        return jsonify({"error": "No hay credenciales configuradas. Ve a Configuración → Google Sheets."}), 400
    if not sheet_url:
        return jsonify({"error": "No hay URL de Google Sheets configurada."}), 400
    if not Path(creds_path).exists():
        return jsonify({"error": f"Archivo de credenciales no encontrado: {creds_path}"}), 400

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds  = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client = gspread.authorize(creds)
        sh     = client.open_by_url(sheet_url)

        # Hoja de Resultados Actuales
        ws = sh.sheet1
        ws.clear()

        # Asegurar que la hoja tenga filas suficientes para evitar errores de límites
        if len(products) + 1 > ws.row_count:
            ws.resize(rows=len(products) + 1)

        HEADERS = ["N", "Fecha/Hora", "Subcategoría", "Marca", "Producto", "SKU", "Precio", "Precio Anterior", "Descuento", "URL", "Presentación", "Tipo Base", "Grupo Competencia"]

        # Colorear cabecera
        ws.update([HEADERS], "A1")
        ws.format("A1:M1", {
            "backgroundColor": {"red": 0.18, "green": 0.42, "blue": 0.71},
            "textFormat": {"bold": True, "foregroundColor": {"red":1,"green":1,"blue":1}, "fontSize": 11},
            "horizontalAlignment": "CENTER",
        })

        # Datos en lotes
        rows = []
        for i, p in enumerate(products, 1):
            name   = p.get("name", "")
            brand  = p.get("brand", "")
            subcat = p.get("subcategory", "")
            pres   = _extract_presentation(name)
            base   = _clean_product_type(name, brand, subcat)
            group  = f"{subcat} | {pres}"

            rows.append([
                i,
                p.get("timestamp", ""),
                subcat,
                brand,
                name,
                p.get("sku", ""),
                p.get("price", ""),
                p.get("oldPrice", ""),
                p.get("discount", ""),
                p.get("link", ""),
                pres,
                base,
                group
            ])

        BATCH = 500
        for start in range(0, len(rows), BATCH):
            batch = rows[start:start+BATCH]
            row_start = start + 2
            ws.update(batch, f"A{row_start}")
        ws.freeze(rows=1)

        # ── Lógica de Histórico ──────────────────────
        try:
            ws_hist = sh.worksheet("Histórico")
        except gspread.exceptions.WorksheetNotFound:
            # Crear con suficientes filas iniciales según el lote actual
            ws_hist = sh.add_worksheet(title="Histórico", rows=str(max(1000, len(products) + 1)), cols=len(HEADERS))
            ws_hist.update([HEADERS], "A1")
            ws_hist.format("A1:M1", {
                "backgroundColor": {"red": 0.1, "green": 0.1, "blue": 0.1},
                "textFormat": {"bold": True, "foregroundColor": {"red":1,"green":1,"blue":1}},
                "horizontalAlignment": "CENTER",
            })
            ws_hist.freeze(rows=1)

        # En el histórico no usamos el índice relativo (i), usamos el total de filas
        hist_rows = []
        for p in products:
            name   = p.get("name", "")
            brand  = p.get("brand", "")
            subcat = p.get("subcategory", "")
            pres   = _extract_presentation(name)
            base   = _clean_product_type(name, brand, subcat)
            group  = f"{subcat} | {pres}"

            hist_rows.append([
                "", # Índice vacío
                p.get("timestamp", ""),
                subcat,
                brand,
                name,
                p.get("sku", ""),
                p.get("price", ""),
                p.get("oldPrice", ""),
                p.get("discount", ""),
                p.get("link", ""),
                pres,
                base,
                group
            ])
        
        ws_hist.append_rows(hist_rows, value_input_option="USER_ENTERED")

        return jsonify({
            "ok":    True,
            "tab":   ws.title,
            "total": len(products),
            "url":   sheet_url,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ── Rutas de IA ────────────────────────────────────────────────────────────

@app.route("/api/ai/config", methods=["GET"])
def get_ai_config():
    cfg = _load_ai_config()
    # No exponer la clave completa al frontend, solo si existe
    return jsonify({"has_key": bool(cfg.get("gemini_api_key", "").strip())})

@app.route("/api/ai/config", methods=["POST"])
def save_ai_config():
    cfg = request.get_json(force=True)
    _save_ai_config({"gemini_api_key": cfg.get("gemini_api_key", "").strip()})
    return jsonify({"ok": True})


@app.route("/api/ai/analyze/<session_id>", methods=["POST"])
def start_ai_analysis(session_id):
    """Inicia el análisis IA de competidores para una sesión de scraping."""
    if session_id not in _sessions:
        return jsonify({"error": "Sesión no encontrada"}), 404

    products = _sessions[session_id].get("products", [])
    if not products:
        return jsonify({"error": "Sin productos para analizar"}), 400

    cfg = _load_ai_config()
    api_key = cfg.get("gemini_api_key", "").strip()
    if not api_key:
        return jsonify({"error": "Configura tu Gemini API Key primero en ⚙️ Configurar IA"}), 400

    ai_sess_id = f"ai_{session_id}"
    _ai_sessions[ai_sess_id] = {
        "queue": Queue(),
        "enriched_products": [],
        "cancelled": False,
        "source_session": session_id,
    }

    t = threading.Thread(
        target=_run_ai_analysis,
        args=(ai_sess_id, products, api_key),
        daemon=True
    )
    t.start()
    return jsonify({"ai_session_id": ai_sess_id, "total": len(products)})


@app.route("/api/ai/events/<ai_session_id>")
def ai_sse_events(ai_session_id):
    """SSE para streaming del progreso del análisis IA."""
    def generate():
        if ai_session_id not in _ai_sessions:
            yield 'data: {"event":"ai_error","data":{"msg":"Sesión IA no encontrada"}}\n\n'
            return
        q: Queue = _ai_sessions[ai_session_id]["queue"]
        while True:
            try:
                msg = q.get(timeout=60)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg["event"] in ("ai_done", "ai_error"):
                    break
            except Empty:
                yield 'data: {"event":"ping"}\n\n'

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/ai/cancel/<ai_session_id>", methods=["POST"])
def cancel_ai_analysis(ai_session_id):
    if ai_session_id in _ai_sessions:
        _ai_sessions[ai_session_id]["cancelled"] = True
    return jsonify({"ok": True})


@app.route("/api/ai/export-sheets/<ai_session_id>", methods=["POST"])
def ai_export_to_sheets(ai_session_id):
    """Exporta resultados del análisis IA a una pestaña 'Competidores IA' en Google Sheets."""
    if ai_session_id not in _ai_sessions:
        return jsonify({"error": "Sesión IA no encontrada"}), 404

    enriched = _ai_sessions[ai_session_id].get("enriched_products", [])
    if not enriched:
        return jsonify({"error": "Sin resultados de IA para exportar"}), 400

    cfg = _load_gsheets_config()
    creds_path = cfg.get("credentials_path", "").strip().strip('"').strip("'")
    sheet_url  = cfg.get("spreadsheet_url", "").strip()

    if not creds_path or not sheet_url:
        return jsonify({"error": "Configura Google Sheets primero"}), 400
    if not Path(creds_path).exists():
        return jsonify({"error": f"Archivo de credenciales no encontrado: {creds_path}"}), 400

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds  = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client = gspread.authorize(creds)
        sh     = client.open_by_url(sheet_url)

        AI_HEADERS = [
            "N", "Fecha/Hora", "Subcategoría", "Marca", "Producto", "SKU",
            "Precio", "Descuento", "URL",
            "🤖 Tipo de Producto (IA)", "🤖 Uso / Para qué sirve", "🤖 Público Objetivo",
            "🤖 Grupo Competidor (IA)", "Presentación (regex)"
        ]

        # Crear o limpiar pestaña
        try:
            ws_ai = sh.worksheet("Competidores IA")
            ws_ai.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws_ai = sh.add_worksheet(
                title="Competidores IA",
                rows=str(max(1000, len(enriched) + 2)),
                cols=len(AI_HEADERS)
            )

        if len(enriched) + 1 > ws_ai.row_count:
            ws_ai.resize(rows=len(enriched) + 2)

        ws_ai.update([AI_HEADERS], "A1")

        # Cabecera estilo IA (morado oscuro)
        col_letter = chr(ord('A') + len(AI_HEADERS) - 1)
        ws_ai.format(f"A1:{col_letter}1", {
            "backgroundColor": {"red": 0.35, "green": 0.10, "blue": 0.60},
            "textFormat": {
                "bold": True,
                "foregroundColor": {"red": 1, "green": 1, "blue": 1},
                "fontSize": 10
            },
            "horizontalAlignment": "CENTER",
            "wrapStrategy": "WRAP",
        })
        ws_ai.freeze(rows=1)

        # Ordenar por grupo competidor para que competidores queden juntos
        enriched_sorted = sorted(
            enriched,
            key=lambda p: (p.get("ai_competitor_group", ""), p.get("brand", ""))
        )

        rows = []
        for i, p in enumerate(enriched_sorted, 1):
            name   = p.get("name", "")
            brand  = p.get("brand", "")
            subcat = p.get("subcategory", "")
            pres   = _extract_presentation(name)
            ai_group = p.get("ai_competitor_group", "") or f"{subcat} | {pres}"

            rows.append([
                i,
                p.get("timestamp", ""),
                subcat,
                brand,
                name,
                p.get("sku", ""),
                p.get("price", ""),
                p.get("discount", ""),
                p.get("link", ""),
                p.get("ai_product_type", ""),
                p.get("ai_use", ""),
                p.get("ai_target", ""),
                ai_group,
                pres,
            ])

        # Colorear filas alternando por grupo competidor
        BATCH = 500
        for start in range(0, len(rows), BATCH):
            batch = rows[start:start + BATCH]
            ws_ai.update(batch, f"A{start + 2}", value_input_option="USER_ENTERED")

        # Resaltar columnas IA con color suave
        ai_col_start = chr(ord('A') + 9)  # columna J
        ai_col_end   = chr(ord('A') + len(AI_HEADERS) - 1)
        ws_ai.format(f"{ai_col_start}2:{ai_col_end}{len(rows)+1}", {
            "backgroundColor": {"red": 0.94, "green": 0.90, "blue": 1.0},
        })

        return jsonify({
            "ok":    True,
            "tab":   "Competidores IA",
            "total": len(enriched),
            "url":   sheet_url,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ── Rutas de Vigilancia de Precios ─────────────────────────────────────────

@app.route("/api/monitor/status", methods=["GET"])
def get_monitor_status():
    return jsonify({
        "active": _monitor_active,
        "alerts_count": len(_load_json(ALERTS_FILE, [])),
        "tracked_count": len(_load_json(WATCHLIST_FILE, []))
    })

@app.route("/api/monitor/toggle/<int:status>", methods=["GET"])
def toggle_monitor(status):
    try:
        global _monitor_active
        active = (status == 1)
        
        # Log de emergencia
        with open("monitor_trigger.txt", "a") as f:
            f.write(f"TRIGGER: {active} at {datetime.now()}\n")

        if active and not _monitor_active:
            _monitor_active = True
            _start_monitor_thread()
        elif not active:
            _monitor_active = False
            
        return jsonify({"active": _monitor_active, "status": "ok"})
    except Exception as e:
        return jsonify({"active": False, "error": str(e)}), 200 # Siempre JSON

@app.route("/api/monitor/watchlist", methods=["POST"])
def sync_watchlist():
    try:
        data = request.get_json(force=True)
        watchlist = data.get("watchlist", [])
        _save_json(WATCHLIST_FILE, watchlist)
        _sync_state_to_sheets() # RESPALDO EN NUBE
        return jsonify({"ok": True, "count": len(watchlist)})
    except:
        return jsonify({"ok": False}), 200

@app.route("/api/monitor/alerts", methods=["GET"])
def get_alerts():
    return jsonify(_load_json(ALERTS_FILE, []))

@app.route("/api/monitor/alerts/clear", methods=["POST"])
def clear_alerts():
    _save_json(ALERTS_FILE, [])
    return jsonify({"ok": True})


@app.route("/api/email/config", methods=["GET"])
def get_email_config():
    return jsonify(_load_email_config())

@app.route("/api/email/config", methods=["POST"])
def save_email_config():
    cfg = request.get_json(force=True)
    _save_json(EMAIL_CONFIG_FILE, cfg)
    _sync_state_to_sheets() # RESPALDO EN NUBE
    return jsonify({"ok": True})

@app.route("/api/email/test", methods=["POST"])
def test_email_notification():
    try:
        watchlist = _load_json(WATCHLIST_FILE, [])
        if not watchlist:
            return jsonify({"ok": False, "error": "Tu Watchlist está vacía. Agrega productos primero."})
        
        # Enviar el primer producto como prueba o un resumen
        prod = watchlist[0]
        success = _send_email_notification(
            product_name=f"[PRUEBA] {prod.get('name', 'Producto')}",
            old_price="10.00",
            new_price=prod.get('price', '0.00'),
            image_url=prod.get('image', ''),
            link=prod.get('link', BASE_URL)
        )
        if success:
            return jsonify({"ok": True})
        else:
            return jsonify({"ok": False, "error": "Error en el servidor SMTP. Revisa tus credenciales."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


if __name__ == "__main__":
    # Recuperar estado de la nube al arrancar
    _recover_state_from_sheets()
    
    print("=" * 55)
    print("  Farmatodo Scraper App - PUERTO 5051")
    print("  MODO: NUBE RESILIENTE (Backup en GSheets)")
    print("=" * 55)
    app.run(host="0.0.0.0", port=5051, debug=False, threaded=True)
